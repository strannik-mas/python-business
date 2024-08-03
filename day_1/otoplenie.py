import openpyxl
import dateutil
def c_to_f(c):
    return (c * 9/5) + 32

f = open(r"d:\курсы\курсы Специалист\Python\[Специалист] Python для бизнес - аналитики (2021)\Code\Датасеты\Отопление.txt", 'r')
lines = f.readlines()[14:44]
wb = openpyxl.Workbook()
ws = wb.active
data = []
for line in lines:
    t1 = float(line[51:56])
    t2 = float(line[57:62])
    t1_f = round(c_to_f(t1), 2)
    t2_f = round(c_to_f(t2), 2)
    d = dateutil.parser.parse(line[1:9], dayfirst=True).date().strftime('%d/%m/%Y')
    row = (d, t1, t2, t1_f, t2_f)
    data.append(row)
for data_row, excel_row in zip(data, ws.iter_rows(min_row=2, max_row=2 + len(data), min_col=1, max_col=1 + len(data[0]))):
    for value, cell in zip(data_row, excel_row):
        cell.value = value
headers = ['Дата', 't1 C', 't2 C', 't1 F', 't2 F']
for name, cell in zip(headers, ws[1]):
    cell.value = name
wb.save('temperature.xlsx')

