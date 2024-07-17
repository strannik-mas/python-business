import openpyxl
a = [12, 22, 11, 55, 232, 222, 6544]
b = [12, 3.14, 11, 55, -1, 222, 0]
wb = openpyxl.Workbook()    #Книга Excel
ws = wb.active #текущий лист
ws['A1'] = 'hello'
ws['B1'] = 'fuck'
for row in ws['A2':'B11']:
    for cell in row:
        # print(cell)
        cell.value = 0

i = 0
for row in ws['C2':'C8']:
    for cell in row:
        cell.value = a[i]
    i += 1

for v, cell in zip(b, ws['D']):
    cell.value = v

i = 1
for cell in ws['E']:
    cell.value = f'=D{i}*2'
    i += 1

ws['C12'] = '=SUM(C2:C11)'

header_font = openpyxl.styles.Font(bold = True, color='800000')
ws['A1'].font = header_font

wb.save('example.xlsx')