import openpyxl

wb = openpyxl.load_workbook(r"d:\курсы\курсы Специалист\Python\[Специалист] Python для бизнес - аналитики (2021)\Code\Датасеты\СМЕТЫ.xlsx")
ws = wb.active
rows_total = ws.max_row
last_section = None
last_subsection = None

def cell_empty(value):
    return not (value is not None and value != 0 and value != '-' and value != '0')

for row_n in range(1, rows_total + 1):
    a_cell = ws.cell(row=row_n, column=1).value
    if type(a_cell) is str and 'раздел' in a_cell.lower() and 'подраздел' not in a_cell.lower():
        last_section = a_cell
        last_subsection = None

    if type(a_cell) is str and 'подраздел' in a_cell.lower():
        last_subsection = a_cell

    # if ws[f'C{row_n}']
    job_text = ws.cell(row=row_n, column=3).value
    if type(job_text) is str and 'бетон' in job_text.lower():
        if cell_empty(ws.cell(row=row_n, column=8).value)\
            and cell_empty(ws.cell(row=row_n, column=9).value)\
            and cell_empty(ws.cell(row=row_n, column=10).value)\
            and cell_empty(ws.cell(row=row_n, column=11).value):
            #материал а не работа
            name = job_text
            volume = ws.cell(row=row_n, column=4).value
            print(name, volume, last_section, last_subsection)

